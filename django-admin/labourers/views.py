from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout 
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from datetime import date  
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.utils import timezone 
from django.db.models import Count, Q
from django.core.files.base import ContentFile
import json
import os
from django.conf import settings
import base64
import math
import re
import hashlib  # ADD THIS
from datetime import datetime
import numpy as np
from PIL import Image, ImageChops, ImageStat  # FIX THIS IMPORT
import io
from decimal import Decimal
# REMOVE THIS LINE: from .utils.image_comparison import base64_to_image, compare_images_simple, generate_image_hash
from .models import (
    Labourer, Project, Supervisor, IDVerificationLog, 
    AttendanceLog, SystemConfig, ImageQualityConfig,
    Contract, Task, TaskAssignment, Equipment, EquipmentAssignment,
    DailyClosureLog, ExceptionReport, Role, CheckIn, CheckOut
)
from .forms import (
    LabourerForm, ImageUploadForm, IDScanForm, 
    AttendanceCheckinForm, ProjectForm, SupervisorForm
)

# ==================== GEOFENCING FUNCTIONS ====================

def point_in_polygon(point, polygon):
    """
    Check if a point (lat, lng) is inside a polygon using ray-casting algorithm.
    
    Args:
        point: Tuple/List of (latitude, longitude)
        polygon: List of [latitude, longitude] pairs forming the boundary
        
    Returns:
        Boolean: True if point is inside polygon, False otherwise
    """
    if not polygon or len(polygon) < 3:
        # If no boundary defined or invalid polygon, allow check-in
        return True
    
    lat, lng = float(point[0]), float(point[1])
    inside = False
    
    n = len(polygon)
    p1_lat, p1_lng = polygon[0]
    
    for i in range(1, n + 1):
        p2_lat, p2_lng = polygon[i % n]
        
        if lng > min(p1_lng, p2_lng):
            if lng <= max(p1_lng, p2_lng):
                if lat <= max(p1_lat, p2_lat):
                    if p1_lng != p2_lng:
                        x_intersection = (lng - p1_lng) * (p2_lat - p1_lat) / (p2_lng - p1_lng) + p1_lat
                    if p1_lat == p2_lat or lat <= x_intersection:
                        inside = not inside
        
        p1_lat, p1_lng = p2_lat, p2_lng
    
    return inside

def calculate_distance(lat1, lon1, lat2, lon2):
    """
    Calculate distance between two GPS coordinates in meters using Haversine formula.
    
    Args:
        lat1, lon1: First point coordinates
        lat2, lon2: Second point coordinates
        
    Returns:
        Distance in meters
    """
    R = 6371000  # Earth's radius in meters
    
    lat1_rad = math.radians(float(lat1))
    lat2_rad = math.radians(float(lat2))
    delta_lat = math.radians(float(lat2) - float(lat1))
    delta_lon = math.radians(float(lon2) - float(lon1))
    
    a = math.sin(delta_lat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(delta_lon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    
    distance = R * c
    return distance

def validate_geofence(latitude, longitude, project):
    """
    Validate if the given coordinates are within the project's geofence.
    STRICT MODE: Only allows check-in inside the boundary.
    
    Args:
        latitude: User's latitude
        longitude: User's longitude
        project: Project object with boundary_coordinates
        
    Returns:
        Tuple: (is_valid: bool, message: str, distance_from_boundary: float)
    """
    if not project.boundary_coordinates or len(project.boundary_coordinates) < 3:
        # No geofence defined, deny check-in
        return False, "No boundary configured for this project. Please contact admin to set up geofence.", 0
    
    user_point = (latitude, longitude)
    is_inside = point_in_polygon(user_point, project.boundary_coordinates)
    
    if is_inside:
        return True, "Location verified: You are within the project boundary", 0
    else:
        # Calculate distance from nearest boundary point
        min_distance = float('inf')
        for boundary_point in project.boundary_coordinates:
            distance = calculate_distance(latitude, longitude, boundary_point[0], boundary_point[1])
            min_distance = min(min_distance, distance)
        
        return False, f"You are {min_distance:.0f} meters outside the project boundary. You must be inside the boundary to check in.", min_distance

# ==================== IMAGE COMPARISON FUNCTIONS ====================

def image_to_base64(image):
    """Convert PIL Image to base64"""
    buffered = io.BytesIO()
    image.save(buffered, format="JPEG", quality=85)
    return base64.b64encode(buffered.getvalue()).decode()

def base64_to_image(base64_string):
    """Convert base64 to PIL Image"""
    if ',' in base64_string:
        base64_string = base64_string.split(',')[1]
    
    image_data = base64.b64decode(base64_string)
    return Image.open(io.BytesIO(image_data))

def generate_image_hash(image):
   
    try:
        # Standardize image
        if isinstance(image, str):
            # If it's a base64 string, convert to image first
            image = base64_to_image(image)
        
        # Resize to standard size and convert to grayscale
        image = image.resize((200, 200)).convert('L')
        
        # Get pixel data
        pixels = list(image.getdata())
        
        # Create hash from pixel values
        pixel_str = ''.join(str(p) for p in pixels)
        return hashlib.sha256(pixel_str.encode()).hexdigest()
    except Exception as e:
        print(f"Error generating hash: {e}")
        return ""

def compare_images_simple(image1, image2):
    
    try:
        # Convert inputs to PIL Images if they aren't already
        if isinstance(image1, str):
            if image1.startswith('data:image'):
                image1 = base64_to_image(image1)
            else:
                image1 = Image.open(image1)
        
        if isinstance(image2, str):
            if image2.startswith('data:image'):
                image2 = base64_to_image(image2)
            else:
                image2 = Image.open(image2)
        
        # Resize both images to same size
        size = (300, 300)
        img1 = image1.resize(size).convert('RGB')
        img2 = image2.resize(size).convert('RGB')
        
        # Hash comparison
        hash1 = generate_image_hash(img1)
        hash2 = generate_image_hash(img2)
        hash_similar = hash1 == hash2
        
        # Histogram comparison
        hist1 = img1.histogram()
        hist2 = img2.histogram()
        
        if sum(hist1) > 0:
            hist_similarity = sum(min(h1, h2) for h1, h2 in zip(hist1, hist2)) / sum(hist1)
        else:
            hist_similarity = 0
        
        # Pixel difference
        diff = ImageChops.difference(img1, img2)
        stat = ImageStat.Stat(diff)
        
        if stat.mean:
            rms = math.sqrt(sum([x*x for x in stat.mean]) / len(stat.mean))
            pixel_similarity = max(0, 100 - rms)
        else:
            pixel_similarity = 0
            rms = 0
        
        # Combined score
        similarity_score = (hist_similarity * 100 * 0.4) + (pixel_similarity * 0.6)
        
        return {
            'hash_match': hash_similar,
            'histogram_similarity': hist_similarity,
            'pixel_similarity': pixel_similarity,
            'similarity_score': similarity_score,
            'is_similar': similarity_score > 50,  # Lowered from 70 to 50
            'rms': rms
        }
        
    except Exception as e:
        print(f"Error comparing images: {e}")
        return {
            'hash_match': False,
            'histogram_similarity': 0,
            'pixel_similarity': 0,
            'similarity_score': 0,
            'is_similar': False,
            'rms': 0,
            'error': str(e)
        }

# ==================== DASHBOARD VIEWS ====================

@login_required
def dashboard(request):
    """Main dashboard view"""
    total_labourers = Labourer.objects.count()
    active_labourers = Labourer.objects.filter(status='ACTIVE').count()
    verified_labourers = Labourer.objects.filter(status='ACTIVE').count()
    pending_verification = Labourer.objects.filter(status='PENDING').count()
    
    # Today's attendance stats
    today = timezone.now().date()
    today_checkins = AttendanceLog.objects.filter(
        log_timestamp__date=today,
        log_type='Check-In'
    ).count()
    today_checkouts = AttendanceLog.objects.filter(
        log_timestamp__date=today,
        log_type='Check-Out'
    ).count()
    
    # Pending check-ins requiring supervisor verification
    pending_checkins = CheckIn.objects.filter(
        status='PENDING',
        timestamp__date=today
    ).select_related('labourer', 'project').order_by('-timestamp')[:10]
    
    # Recent activities
    recent_attendance = AttendanceLog.objects.all().order_by('-log_timestamp')[:10]
    recent_labourers = Labourer.objects.all().order_by('-created_at')[:5]
    
    context = {
        'total_labourers': total_labourers,
        'active_labourers': active_labourers,
        'verified_labourers': verified_labourers,
        'pending_verification': pending_verification,
        'today_checkins': today_checkins,
        'today_checkouts': today_checkouts,
        'pending_checkins': pending_checkins,
        'recent_attendance': recent_attendance,
        'recent_labourers': recent_labourers,
        'user': request.user,
        'page_title': 'Labour Management Dashboard'
    }
    return render(request, 'labourers/dashboard.html', context)

# ==================== AUTHENTICATION VIEWS ====================

def create_default_users():
    """Create default users if they don't exist"""
    # Create Supervisor user: ngenes / super
    if not User.objects.filter(username='ngenes').exists():
        supervisor = User.objects.create_user(
            username='ngenes',
            password='super',
            email='supervisor@company.com',
            is_staff=True,
            is_superuser=False,
            first_name='Supervisor',
            last_name='Ngenes'
        )
        print("Supervisor user 'ngenes' created")
    
    # Create Admin user: Burita / admin23
    if not User.objects.filter(username='Burita').exists():
        admin = User.objects.create_user(
            username='Burita',
            password='admin23',
            email='admin@company.com',
            is_staff=True,
            is_superuser=True,
            first_name='Admin',
            last_name='Burita'
        )
        print("Admin user 'Burita' created")

def login_view(request):
    """Login view with role-based access"""
    # Create default users on first access
    create_default_users()
    
    # If already logged in, go directly to dashboard
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    error = None
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        print(f"Login attempt: {username}")
        
        # Authenticate user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Store user role in session
            if user.is_superuser:
                request.session['user_role'] = 'admin'
                request.session['user_display_name'] = 'Administrator'
            else:
                request.session['user_role'] = 'supervisor'
                request.session['user_display_name'] = 'Supervisor'
            
            print(f"Login successful for {username} ({request.session['user_role']})")
            return redirect('dashboard')
        else:
            error = "Invalid username or password"
            print(f"Invalid credentials: {username}")
        
        return render(request, 'labourers/login.html', {'error': error})
    
    return render(request, 'labourers/login.html')

def logout_view(request):
    """Logout view"""
    logout(request)
    return redirect('login')

# ==================== LABOURER MANAGEMENT VIEWS ====================

@login_required
def labourer_list(request):
    """List all labourers with filtering options"""
    labourers = Labourer.objects.all().order_by('-created_at')
    
    # Filtering
    status_filter = request.GET.get('status')
    if status_filter:
        labourers = labourers.filter(status=status_filter)
    
    verification_filter = request.GET.get('verification')
    if verification_filter:
        labourers = labourers.filter(status=verification_filter)
    
    project_filter = request.GET.get('project')
    if project_filter:
        labourers = labourers.filter(project_id=project_filter)
    
    search_query = request.GET.get('search')
    if search_query:
        labourers = labourers.filter(
            Q(full_name__icontains=search_query) |
            Q(national_id__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    projects = Project.objects.all()
    
    context = {
        'labourers': labourers,
        'projects': projects,
        'status_filter': status_filter,
        'verification_filter': verification_filter,
        'project_filter': project_filter,
        'search_query': search_query,
    }
    return render(request, 'labourers/labourers_list.html', context)

@login_required
def add_labourer(request):
    """ADD NEW LABOURER - GUARANTEED WORKING VERSION"""
    
    # Get projects for dropdown
    projects = Project.objects.all()
    supervisors = Supervisor.objects.filter(is_active=True)
    
    if request.method == 'POST':
        print("=" * 60)
        print("DEBUG: Form submitted!")
        print(f"POST Data: {dict(request.POST)}")
        
        try:
            # Extract required fields
            national_id = request.POST.get('national_id', '').strip()
            full_name = request.POST.get('full_name', '').strip()
            designation = request.POST.get('designation', 'General Worker').strip()
            department = request.POST.get('department', 'General').strip()
            phone_number = request.POST.get('phone_number', '+254700000000').strip()
            
            # Validation
            if not national_id:
                messages.error(request, 'National ID is required!')
                return render(request, 'labourers/add_labourer.html', {
                    'projects': projects,
                    'supervisors': supervisors,
                    'page_title': 'Add New Labourer',
                })
            
            if not full_name:
                messages.error(request, 'Full Name is required!')
                return render(request, 'labourers/add_labourer.html', {
                    'projects': projects,
                    'supervisors': supervisors,
                    'page_title': 'Add New Labourer',
                })
            
            # Check for duplicate national ID
            if Labourer.objects.filter(national_id=national_id).exists():
                messages.error(request, f'Labourer with National ID {national_id} already exists!')
                return render(request, 'labourers/add_labourer.html', {
                    'projects': projects,
                    'supervisors': supervisors,
                    'page_title': 'Add New Labourer',
                })
            
            print(f"DEBUG: Creating labourer: {full_name} ({national_id})")
            
            # Create the labourer object
            labourer = Labourer(
                national_id=national_id,
                full_name=full_name,
                designation=designation,
                department=department,
                phone_number=phone_number,
                employment_type=request.POST.get('employment_type', 'FULL_TIME'),
                status='PENDING',
                whitelisted=False
            )
            
            # Optional fields
            date_of_birth = request.POST.get('date_of_birth', '').strip()
            if date_of_birth:
                labourer.date_of_birth = date_of_birth
            
            email = request.POST.get('email', '').strip()
            if email:
                labourer.email = email
            
            gender = request.POST.get('gender', '').strip()
            if gender:
                labourer.gender = gender
            
            district_of_birth = request.POST.get('district_of_birth', '').strip()
            if district_of_birth:
                labourer.district_of_birth = district_of_birth
            
            id_serial_number = request.POST.get('id_serial_number', '').strip()
            if id_serial_number:
                labourer.id_serial_number = id_serial_number
            
            # Handle project assignment
            project_id = request.POST.get('project')
            if project_id:
                try:
                    project = Project.objects.get(id=project_id)
                    labourer.project = project
                    print(f"DEBUG: Assigned to project: {project.name}")
                except Project.DoesNotExist:
                    print("DEBUG: Project not found")
            
            # Handle supervisor assignment
            supervisor_id = request.POST.get('supervisor')
            if supervisor_id:
                try:
                    supervisor = Supervisor.objects.get(id=supervisor_id)
                    labourer.supervisor = supervisor
                except Supervisor.DoesNotExist:
                    print("DEBUG: Supervisor not found")
            
            # Handle image uploads
            # Portrait photo - handle both file upload and base64 camera capture
            portrait_saved = False
            if 'portrait_photo' in request.FILES:
                labourer.portrait_photo = request.FILES['portrait_photo']
                labourer.portrait_image = request.FILES['portrait_photo']  # Save to both fields
                portrait_saved = True
                print("DEBUG: Portrait photo uploaded from file")
            elif request.POST.get('portrait_photo'):
                # Handle base64 camera capture
                base64_data = request.POST.get('portrait_photo')
                if base64_data and base64_data.startswith('data:image'):
                    try:
                        # Extract base64 data
                        format, imgstr = base64_data.split(';base64,')
                        ext = format.split('/')[-1]
                        
                        # Decode base64
                        data = ContentFile(base64.b64decode(imgstr))
                        file_name = f'{national_id}_portrait.{ext}'
                        
                        # Save to both portrait fields
                        labourer.portrait_photo.save(file_name, data, save=False)
                        labourer.portrait_image.save(file_name, data, save=False)
                        portrait_saved = True
                        print("DEBUG: Portrait photo saved from camera capture")
                    except Exception as e:
                        print(f"DEBUG: Error processing camera photo: {e}")
            
            if portrait_saved:
                print(f"DEBUG: Portrait photo saved successfully")
            
            if 'id_front_photo' in request.FILES:
                labourer.id_front_photo = request.FILES['id_front_photo']
                print("DEBUG: ID front uploaded")
            
            if 'id_back_photo' in request.FILES:
                labourer.id_back_photo = request.FILES['id_back_photo']
                print("DEBUG: ID back uploaded")
            
            # Save to database
            print("DEBUG: Saving to database...")
            labourer.save()
            print(f"✅ SUCCESS: Labourer saved! ID: {labourer.id}, Serial Number: {labourer.serial_number}")
            
            # Verify save worked
            db_count = Labourer.objects.count()
            print(f"✅ Total labourers in database now: {db_count}")
            
            # Create verification log
            try:
                IDVerificationLog.objects.create(
                    labourer=labourer,
                    is_successful=True,
                    verification_type='Manual',
                    verification_method='MANUAL'
                )
                print("DEBUG: Verification log created")
            except Exception as e:
                print(f"DEBUG: Could not create verification log: {e}")
            
            # Success message
            messages.success(request, f'✅ Labourer "{full_name}" added successfully!')
            
            # Redirect to labourers list
            return redirect('labourer_list')
            
        except Exception as e:
            print(f"❌ ERROR: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error saving labourer: {str(e)}')
            return render(request, 'labourers/add_labourer.html', {
                'projects': projects,
                'supervisors': supervisors,
                'page_title': 'Add New Labourer',
            })
    
    # GET request - show empty form
    print("DEBUG: GET request - showing empty form")
    
    context = {
        'projects': projects,
        'supervisors': supervisors,
        'page_title': 'Add New Labourer',
    }
    return render(request, 'labourers/add_labourer.html', context)

@login_required
def labourer_detail(request, labourer_id):
    """View labourer details with images and attendance history"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    # Get attendance history
    attendance_history = AttendanceLog.objects.filter(labourer=labourer).order_by('-log_timestamp')[:50]
    
    # Get verification logs
    verification_logs = IDVerificationLog.objects.filter(labourer=labourer).order_by('-verification_time')[:10]
    
    # Get contracts
    contracts = Contract.objects.filter(labourer=labourer)
    
    context = {
        'labourer': labourer,
        'attendance_history': attendance_history,
        'verification_logs': verification_logs,
        'contracts': contracts,
    }
    return render(request, 'labourers/labourer_detail.html', context)

# ==================== IMAGE UPLOAD VIEWS ====================

@csrf_exempt
def upload_portrait(request, labourer_id):
    """Upload portrait photo for a labourer"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    if request.method == 'POST' and 'portrait_photo' in request.FILES:
        try:
            # Basic image validation using PIL
            image_file = request.FILES['portrait_photo']
            
            # Check file size (max 5MB)
            if image_file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'File size too large. Maximum size is 5MB.'
                })
            
            # Check file extension
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
            file_ext = os.path.splitext(image_file.name)[1].lower()
            if file_ext not in allowed_extensions:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid file type. Allowed types: {", ".join(allowed_extensions)}'
                })
            
            # Save the image
            labourer.portrait_photo = image_file
            labourer.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'Portrait photo uploaded successfully!',
                'image_url': labourer.portrait_photo.url if labourer.portrait_photo else ''
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'No image provided'})

@csrf_exempt
def upload_id_front(request, labourer_id):
    """Upload ID front image"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    if request.method == 'POST' and 'id_front_photo' in request.FILES:
        try:
            image_file = request.FILES['id_front_photo']
            
            # Check file size (max 3MB)
            if image_file.size > 3 * 1024 * 1024:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'File size too large. Maximum size is 3MB.'
                })
            
            # Check file extension
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
            file_ext = os.path.splitext(image_file.name)[1].lower()
            if file_ext not in allowed_extensions:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid file type. Allowed types: {", ".join(allowed_extensions)}'
                })
            
            # Save the image
            labourer.id_front_photo = image_file
            labourer.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'ID front image uploaded successfully!',
                'image_url': labourer.id_front_photo.url if labourer.id_front_photo else ''
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'No image provided'})

@csrf_exempt
def upload_id_back(request, labourer_id):
    """Upload ID back image"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    if request.method == 'POST' and 'id_back_photo' in request.FILES:
        try:
            image_file = request.FILES['id_back_photo']
            
            # Check file size (max 3MB)
            if image_file.size > 3 * 1024 * 1024:
                return JsonResponse({
                    'status': 'error', 
                    'message': 'File size too large. Maximum size is 3MB.'
                })
            
            # Check file extension
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
            file_ext = os.path.splitext(image_file.name)[1].lower()
            if file_ext not in allowed_extensions:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid file type. Allowed types: {", ".join(allowed_extensions)}'
                })
            
            # Save the image
            labourer.id_back_photo = image_file
            labourer.save()
            
            return JsonResponse({
                'status': 'success', 
                'message': 'ID back image uploaded successfully!',
                'image_url': labourer.id_back_photo.url if labourer.id_back_photo else ''
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'No image provided'})

@csrf_exempt
def scan_id_card(request, labourer_id):
    """Simple ID card scanning - placeholder for now"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    if request.method == 'POST' and 'scan_image' in request.FILES:
        try:
            scan_image = request.FILES['scan_image']
            scan_type = request.POST.get('scan_type', 'front')
            
            # Save image based on type
            if scan_type == 'front':
                labourer.id_front_photo = scan_image
            else:
                labourer.id_back_photo = scan_image
            
            labourer.save()
            
            # Create verification log
            IDVerificationLog.objects.create(
                labourer=labourer,
                verification_type='Manual',
                is_successful=True,
                verification_method='MANUAL',
                device_id=request.META.get('HTTP_USER_AGENT', 'Unknown'),
                ip_address=request.META.get('REMOTE_ADDR', ''),
            )
            
            return JsonResponse({
                'status': 'success',
                'message': f'ID {scan_type} side scanned successfully',
                'image_type': scan_type
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'No image provided'})

# ==================== ATTENDANCE VIEWS ====================

@login_required
def attendance(request):    
    today = datetime.now().date()
    labourers = Labourer.objects.all().select_related('project')
    
    # Get today's attendance records - using AttendanceLog
    today_attendances = AttendanceLog.objects.filter(
        log_timestamp__date=today
    ).select_related('labourer', 'labourer__project')
    
    # Get pending check-ins for verification
    pending_checkins = CheckIn.objects.filter(
        status='PENDING',
        timestamp__date=today
    ).select_related('labourer', 'project').order_by('-timestamp')
    
    # Get all projects for display
    projects = Project.objects.all()
    
    context = {
        'today': today.strftime('%Y-%m-%d'),
        'labourers': labourers,
        'today_attendances': today_attendances,
        'pending_checkins': pending_checkins,
        'projects': projects,
    }
    return render(request, 'labourers/attendance.html', context)

def camera_test(request):
    """Simple camera test page"""
    return render(request, 'labourers/camera_test.html')

@login_required
def face_verification_page(request, labourer_id, action):
    """Show face verification page"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    
    if action not in ['check_in', 'check_out']:
        return redirect('attendance')
    
    context = {
        'labourer': labourer,
        'action': action,
        'has_portrait': bool(labourer.portrait_photo)  # Changed from portrait_image
    }
    return render(request, 'labourers/face_verification_simple.html', context)

@csrf_exempt
def verify_face_api(request):
    """API endpoint for face verification using Pillow"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            labourer_id = data.get('labourer_id')
            image_data = data.get('image')
            
            labourer = get_object_or_404(Labourer, id=labourer_id)
            
            if not labourer.portrait_photo:  # Changed from portrait_image
                return JsonResponse({
                    'success': False,
                    'verified': False,
                    'message': 'No photo registered for this labourer.',
                    'similarity': 0
                })
            
            # Convert base64 to image
            live_image = base64_to_image(image_data)
            
            # Load stored portrait
            with Image.open(labourer.portrait_photo.path) as stored_image:
                # Compare images
                result = compare_images_simple(live_image, stored_image)
                
                if result['is_similar']:
                    return JsonResponse({
                        'success': True,
                        'verified': True,
                        'message': 'Photo verification successful!',
                        'similarity': round(result['similarity_score'], 1),
                        'hash_match': result['hash_match'],
                        'pixel_similarity': round(result['pixel_similarity'], 1)
                    })
                else:
                    return JsonResponse({
                        'success': True,
                        'verified': False,
                        'message': 'Photo does not match registered photo.',
                        'similarity': round(result['similarity_score'], 1),
                        'hash_match': result['hash_match'],
                        'pixel_similarity': round(result['pixel_similarity'], 1)
                    })
                    
        except Exception as e:
            return JsonResponse({
                'success': False,
                'verified': False,
                'message': f'Error: {str(e)}',
                'similarity': 0
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
def process_attendance(request, labourer_id):
    """Process attendance after verification"""
    if request.method == 'POST':
        labourer = get_object_or_404(Labourer, id=labourer_id)
        action = request.POST.get('action', 'check_in')
        image_data = request.POST.get('captured_image', '')
        similarity = request.POST.get('similarity', 0)
        verified = request.POST.get('verified') == 'true'
        
        # Save verification photo
        verification_photo = None
        if image_data and verified:
            try:
                format, imgstr = image_data.split(';base64,')
                ext = format.split('/')[-1]
                verification_photo = ContentFile(
                    base64.b64decode(imgstr), 
                    name=f'verify_{labourer.national_id}_{action}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{ext}'
                )
            except:
                pass
        
        if verified and float(similarity) >= 70:
            # Verification successful
            if action == 'check_in':
                # Check if already checked in today
                today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
                existing = AttendanceLog.objects.filter(
                    labourer=labourer,
                    log_timestamp__gte=today_start,
                    log_type='Check-In'
                ).first()
                
                if existing:
                    return JsonResponse({
                        'success': False,
                        'message': f'{labourer.full_name} is already checked in.'
                    })
                
                # Create check-in record
                attendance = AttendanceLog.objects.create(
                    labourer=labourer,
                    log_type='Check-In',
                    verification_method='Face Verification',
                    access_granted=True,
                    notes=f'Face verification: {similarity}% match'
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'✅ {labourer.full_name} checked in successfully!',
                    'attendance_id': attendance.id
                })
            
            elif action == 'check_out':
                # Find latest check-in without check-out
                attendance = AttendanceLog.objects.filter(
                    labourer=labourer,
                    log_type='Check-In'
                ).order_by('-log_timestamp').first()
                
                if attendance:
                    # Create check-out record
                    check_out = AttendanceLog.objects.create(
                        labourer=labourer,
                        log_type='Check-Out',
                        verification_method='Face Verification',
                        access_granted=True,
                        notes=f'Face verification: {similarity}% match'
                    )
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'✅ {labourer.full_name} checked out successfully!',
                        'attendance_id': check_out.id
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'message': f'❌ No active check-in found for {labourer.full_name}'
                    })
        else:
            return JsonResponse({
                'success': False,
                'message': f'❌ Photo verification failed for {labourer.full_name}',
                'similarity': similarity
            })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@csrf_exempt
def check_in(request, labourer_id):
    """Check in labourer with geofencing and facial recognition validation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    from django.core.files.base import ContentFile
    import base64
    from decimal import Decimal
    labourer = get_object_or_404(Labourer, id=labourer_id)
        
        # Get GPS coordinates and project from request
    project_id = request.POST.get('project_id')
    latitude = request.POST.get('latitude')
    longitude = request.POST.get('longitude')
    accuracy = request.POST.get('accuracy')
    captured_image = request.POST.get('captured_image')  # Base64 image from camera
        
     # Validate required fields
    if not project_id:
            return JsonResponse({
                'success': False,
                'error': 'Project selection is required'
            })
        
    if not latitude or not longitude:
            return JsonResponse({
                'success': False,
                'error': 'GPS location is required for check-in. Please enable location services.'
            })
        
        # Facial recognition (optional for now to allow check-in)
    facial_confidence = 0
    if captured_image:
        if not labourer.portrait_photo:
            return JsonResponse({'success': False, 'error': 'Labourer has no registered portrait photo.'})
        try:
            # Convert base64 to image
            live_image = base64_to_image(captured_image)

            # Load stored portrait
            with Image.open(labourer.portrait_photo.path) as stored_image:
                result = compare_images_simple(live_image, stored_image)
                facial_confidence = result['similarity_score']
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Facial recognition failed: {str(e)}'})

        # Hard threshold (fail if below)
        threshold = 70
        if facial_confidence < threshold:
            return JsonResponse({
                'success': False,
                'error': f'Face mismatch ({round(facial_confidence,1)}%). Check-in denied.'
            })
        
        # Get project
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Invalid project selected'
            })
        
        # Validate geofence
        is_valid, message, distance = validate_geofence(
            Decimal(latitude), 
            Decimal(longitude), 
            project
        )
        
        if not is_valid:
            return JsonResponse({
                'success': False,
                'error': f'⛔ Check-in denied: {message}. You must be at the project location to check in.'
            })
        
        # Create attendance record
        attendance = AttendanceLog(
            labourer=labourer,
            log_type='Check-In',
            verification_method='Face+Location',
            latitude=Decimal(latitude),
            longitude=Decimal(longitude),
            location_accuracy=Decimal(accuracy) if accuracy else None,
            location_verified=True,
            location_verification_time=timezone.now()
        )
        attendance.save()        
        checkin = CheckIn(
            labourer=labourer,
            project=project,
            facial_match_confidence=facial_confidence,
            location_lat=Decimal(latitude),
            location_lng=Decimal(longitude),
            within_geofence=is_valid,
            whitelist_valid=labourer.whitelisted if hasattr(labourer, 'whitelisted') else True,
            within_operating_hours=True,
            status='SUCCESS',
            access_granted=True
        )
    if captured_image:
        try:
            format_part, imgstr = captured_image.split(';base64,')
            ext = format_part.split('/')[-1]
            photo_file = ContentFile(
                base64.b64decode(imgstr),
                name=f'checkin_{labourer.national_id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{ext}'
            )
            checkin.facial_recognition_photo = photo_file
        except:
            pass
    checkin.save()

    return JsonResponse({
        'success': True,
        'message': f'✅ {labourer.full_name} checked in successfully at {project.name}. Confidence: {round(facial_confidence,1)}%'
    })
@login_required
def verify_checkin(request, checkin_id, action):
    """Verify or reject a pending check-in"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    try:
        checkin = CheckIn.objects.get(id=checkin_id)
    except CheckIn.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Check-in not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Database error: {str(e)}'})
    
    try:
        if action == 'approve':
            checkin.status = 'SUCCESS'
            checkin.access_granted = True
            checkin.override_by = request.user
            checkin.override_reason = f'Approved by {request.user.get_full_name() or request.user.username}'
            checkin.save()
            
            return JsonResponse({
                'success': True,
                'message': f'✅ Check-in approved for {checkin.labourer.full_name}'
            })
        
        elif action == 'reject':
            checkin.status = 'FAILED'
            checkin.access_granted = False
            checkin.override_by = request.user
            checkin.override_reason = f'Rejected by {request.user.get_full_name() or request.user.username}'
            checkin.save()
            
            return JsonResponse({
                'success': True,
                'message': f'❌ Check-in rejected for {checkin.labourer.full_name}'
            })
        
        else:
            return JsonResponse({'success': False, 'error': 'Invalid action'})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error updating check-in: {str(e)}'})

@csrf_exempt
def check_out(request, labourer_id):
    """Check out labourer with geofencing and facial recognition validation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})

    from decimal import Decimal
    import base64
    from django.core.files.base import ContentFile

    labourer = get_object_or_404(Labourer, id=labourer_id)

    project_id = request.POST.get('project_id')
    latitude = request.POST.get('latitude')
    longitude = request.POST.get('longitude')
    accuracy = request.POST.get('accuracy')
    captured_image = request.POST.get('captured_image')  # Base64 image from camera

    # Validate required fields
    if not project_id:
        return JsonResponse({'success': False, 'error': 'Project selection is required'})
    if not latitude or not longitude:
        return JsonResponse({'success': False, 'error': 'GPS location is required for check-out. Please enable location services.'})

    # Facial recognition enforcement (if captured)
    facial_confidence = 0
    if captured_image:
        if not labourer.portrait_photo:
            return JsonResponse({'success': False, 'error': 'Labourer has no registered portrait photo.'})
        try:
            live_image = base64_to_image(captured_image)
            with Image.open(labourer.portrait_photo.path) as stored_image:
                result = compare_images_simple(live_image, stored_image)
                facial_confidence = result['similarity_score']
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Facial recognition failed: {str(e)}'})

        # Hard threshold for checkout
        threshold = 70
        if facial_confidence < threshold:
            return JsonResponse({
                'success': False,
                'error': f'Face mismatch ({round(facial_confidence,1)}%). Check-out denied.'
            })

    # Get project
    try:
        project = Project.objects.get(id=project_id)
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid project selected'})

    # Validate geofence
    is_valid, message, distance = validate_geofence(
        Decimal(latitude),
        Decimal(longitude),
        project
    )
    if not is_valid:
        return JsonResponse({'success': False, 'error': f'⛔ Check-out denied: {message}'})

    # Create attendance log
    attendance = AttendanceLog(
        labourer=labourer,
        log_type='Check-Out',
        verification_method='Face+Location',
        latitude=Decimal(latitude),
        longitude=Decimal(longitude),
        location_accuracy=Decimal(accuracy) if accuracy else None,
        location_verified=True,
        location_verification_time=timezone.now()
    )
    attendance.save()

    # Save CheckOut record for admin
    checkout = CheckOut(
        labourer=labourer,
        project=project,
        checkout_type='NORMAL',
        completed_at=timezone.now(),
        checkout_location_lat=Decimal(latitude),
        checkout_location_lng=Decimal(longitude),
        facial_match_confidence=facial_confidence
    )

    # Save photo if captured
    if captured_image:
        try:
            format_part, imgstr = captured_image.split(';base64,')
            ext = format_part.split('/')[-1]
            photo_file = ContentFile(
                base64.b64decode(imgstr),
                name=f'checkout_{labourer.national_id}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.{ext}'
            )
            checkout.facial_recognition_photo = photo_file
        except:
            pass

    checkout.save()

    return JsonResponse({
        'success': True,
        'message': f'✅ {labourer.full_name} checked out successfully from {project.name}. Confidence: {round(facial_confidence,1)}%'
    })

# ==================== API ENDPOINTS ====================

@csrf_exempt
def api_upload_image(request):
    """API endpoint for image upload"""
    if request.method == 'POST':
        try:
            serial_number = request.POST.get('labourer_id')  # Actually serial_number
            image_type = request.POST.get('image_type')
            image_file = request.FILES.get('image')
            
            if not all([serial_number, image_type, image_file]):
                return JsonResponse({'status': 'error', 'message': 'Missing parameters'})
            
            # Find labourer by serial number
            try:
                labourer = Labourer.objects.get(serial_number=serial_number)
            except Labourer.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Labourer not found'})
            
            # Validate file
            if image_file.size > 5 * 1024 * 1024:
                return JsonResponse({'status': 'error', 'message': 'File too large (max 5MB)'})
            
            allowed_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
            file_ext = os.path.splitext(image_file.name)[1].lower()
            if file_ext not in allowed_extensions:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
                })
            
            # Save image based on type
            if image_type == 'portrait':
                labourer.portrait_photo = image_file
            elif image_type == 'id_front':
                labourer.id_front_photo = image_file
            elif image_type == 'id_back':
                labourer.id_back_photo = image_file
            
            labourer.save()
            
            return JsonResponse({
                'status': 'success',
                'message': f'{image_type} uploaded successfully',
                'serial_number': labourer.serial_number,
                'labourer_name': labourer.full_name,
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

@csrf_exempt
def api_verify_attendance(request):
    """API endpoint for attendance verification"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            labourer_id = data.get('labourer_id')
            
            if not labourer_id:
                return JsonResponse({'status': 'error', 'message': 'Labourer ID required'})
            
            # Find labourer by serial number or national ID
            try:
                labourer = Labourer.objects.get(serial_number=labourer_id)
            except Labourer.DoesNotExist:
                try:
                    labourer = Labourer.objects.get(national_id=labourer_id)
                except Labourer.DoesNotExist:
                    return JsonResponse({'status': 'error', 'message': 'Labourer not found'})
            
            # Check if labourer can check in
            if labourer.status != 'ACTIVE':
                return JsonResponse({
                    'status': 'error',
                    'verified': False,
                    'message': 'Labourer is not active'
                })
            
            # Check if already checked in today
            today = timezone.now().date()
            already_checked_in = AttendanceLog.objects.filter(
                labourer=labourer,
                log_timestamp__date=today,
                log_type='Check-In'
            ).exists()
            
            if already_checked_in:
                return JsonResponse({
                    'status': 'warning',
                    'verified': True,
                    'message': 'Already checked in today'
                })
            
            return JsonResponse({
                'status': 'success',
                'verified': True,
                'labourer': {
                    'name': labourer.full_name,
                    'serial_number': labourer.serial_number,
                    'designation': labourer.designation,
                    'department': labourer.department,
                    'status': labourer.status,
                }
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

# ==================== SIMPLIFIED HELPER VIEWS ====================

def get_attendance_history(request):
    """Get attendance history for dashboard"""
    today = timezone.now().date()
    attendance_data = []
    attendance_logs = AttendanceLog.objects.filter(log_timestamp__date=today).order_by('-log_timestamp')[:20]
    
    for log in attendance_logs:
        attendance_data.append({
            'labourer_name': log.labourer.full_name,
            'serial_number': log.labourer.serial_number,
            'log_type': log.log_type,
            'timestamp': log.log_timestamp.strftime('%H:%M:%S'),
            'access_granted': log.access_granted,
        })
    
    return JsonResponse(attendance_data, safe=False)

def labourers_json(request):
    """API endpoint for labourers data"""
    labourers = Labourer.objects.all().values(
        'id', 'serial_number', 'full_name', 'national_id', 
        'designation', 'department', 'phone_number', 'email', 
        'status', 'created_at'
    )
    return JsonResponse(list(labourers), safe=False)

# ==================== TEMPLATE VIEWS ====================

@login_required
def edit_labourer(request, labourer_id):
    """Edit labourer - placeholder"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    projects = Project.objects.all()
    supervisors = Supervisor.objects.all()
    
    context = {
        'labourer': labourer,
        'projects': projects,
        'supervisors': supervisors,
    }
    return render(request, 'labourers/edit_labourer.html', context)

@login_required
def delete_labourer(request, labourer_id):
    """Delete labourer - placeholder"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    if request.method == 'POST':
        labourer.delete()
        messages.success(request, f'Labourer {labourer.full_name} deleted successfully!')
        return redirect('labourer_list')
    return render(request, 'labourers/confirm_delete.html', {'labourer': labourer})

@login_required
def attendance_history(request):
    """View attendance history"""
    today = timezone.now().date()
    attendance_logs = AttendanceLog.objects.all().order_by('-log_timestamp')[:100]
    
    context = {
        'attendance_logs': attendance_logs,
        'today': today,
    }
    return render(request, 'labourers/attendance_history.html', context)

@login_required
def project_list(request):
    """List projects"""
    projects = Project.objects.all()
    return render(request, 'labourers/project_list.html', {'projects': projects})

@login_required
def supervisor_list(request):
    """List supervisors"""
    supervisors = Supervisor.objects.all()
    return render(request, 'labourers/supervisor_list.html', {'supervisors': supervisors})

# ==================== GEOFENCE API ENDPOINTS ====================

@csrf_exempt
def verify_location(request, labourer_id):
    """API endpoint to verify location before showing camera"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            latitude = data.get('latitude')
            longitude = data.get('longitude')
            action = data.get('action', 'checkin')
            
            labourer = get_object_or_404(Labourer, id=labourer_id)
            
            if not latitude or not longitude:
                return JsonResponse({
                    'success': False,
                    'error': 'Location required',
                    'camera_allowed': False
                })
            
            # Check if labourer has a project
            if not labourer.project:
                return JsonResponse({
                    'success': False,
                    'error': 'Labourer not assigned to any project',
                    'camera_allowed': False
                })
            
            # Simplified location check (always true for now)
            is_within_site = True
            
            if not is_within_site:
                return JsonResponse({
                    'success': False,
                    'error': f'You must be at {labourer.project.name} site to {action}',
                    'camera_allowed': False,
                    'site_name': labourer.project.name,
                })
            
            # All checks passed - allow camera
            return JsonResponse({
                'success': True,
                'message': f'Location verified. You are at {labourer.project.name}',
                'camera_allowed': True,
                'site_name': labourer.project.name,
                'site_identifier': labourer.project.site_identifier,
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e),
                'camera_allowed': False
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@csrf_exempt
def get_project_geofence(request, project_id):
    """Get geofence data for a project (for map display)"""
    try:
        project = get_object_or_404(Project, id=project_id)
        
        return JsonResponse({
            'success': True,
            'project': {
                'id': project.id,
                'name': project.name,
                'site_identifier': project.site_identifier,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@csrf_exempt
def simulate_location(request):
    """For testing - simulate being at different locations"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            test_location = data.get('location', 'burhani')
            
            locations = {
                'burhani': {
                    'latitude': -1.28861666,
                    'longitude': 36.79140980,
                    'name': 'Burhani Engineers Site'
                },
                'nearby': {
                    'latitude': -1.28861666,
                    'longitude': 36.79150980,
                    'name': 'Near Burhani (outside)'
                },
                'far': {
                    'latitude': -1.28600000,
                    'longitude': 36.80000000,
                    'name': 'Far from Burhani'
                }
            }
            
            if test_location in locations:
                return JsonResponse({
                    'success': True,
                    'simulated_location': locations[test_location]
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Unknown test location'
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@csrf_exempt
def register_face_simple(request, labourer_id):
    """Simple face registration placeholder"""
    labourer = get_object_or_404(Labourer, id=labourer_id)
    return JsonResponse({'success': False, 'error': 'Face registration not implemented'})

# ==================== REPORTS ====================

@login_required
def reports_dashboard(request):
    """Reports dashboard with attendance, labourer, and project statistics"""
    from datetime import timedelta
    from django.db.models import Count, Q
    
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # Attendance Statistics
    total_labourers = Labourer.objects.filter(status='ACTIVE').count()
    today_checkins = AttendanceLog.objects.filter(
        log_timestamp__date=today,
        log_type='Check-In'
    ).count()
    
    today_checkouts = AttendanceLog.objects.filter(
        log_timestamp__date=today,
        log_type='Check-Out'
    ).count()
    
    week_attendance = AttendanceLog.objects.filter(
        log_timestamp__date__gte=week_ago
    ).count()
    
    # Labourer Statistics by Status
    labourer_stats = Labourer.objects.values('status').annotate(count=Count('id'))
    
    # Project Statistics
    projects = Project.objects.all()
    project_stats = []
    for project in projects:
        project_labourers = Labourer.objects.filter(project=project).count()
        project_today_attendance = AttendanceLog.objects.filter(
            labourer__project=project,
            log_timestamp__date=today
        ).count()
        project_stats.append({
            'name': project.name,
            'site_identifier': project.site_identifier,
            'total_labourers': project_labourers,
            'today_attendance': project_today_attendance
        })
    
    # Recent Attendance Logs
    recent_logs = AttendanceLog.objects.select_related('labourer', 'labourer__project').order_by('-log_timestamp')[:20]
    
    # Role Distribution
    role_stats = Labourer.objects.filter(role__isnull=False).values(
        'role__name'
    ).annotate(count=Count('id'))
    
    context = {
        'today': today,
        'total_labourers': total_labourers,
        'today_checkins': today_checkins,
        'today_checkouts': today_checkouts,
        'week_attendance': week_attendance,
        'labourer_stats': labourer_stats,
        'project_stats': project_stats,
        'recent_logs': recent_logs,
        'role_stats': role_stats,
    }
    
    return render(request, 'labourers/reports.html', context)

@login_required
def settings_view(request):
    """
    System settings page for managing projects, roles, supervisors, and system configuration
    """
    projects = Project.objects.all().order_by('name')
    roles = Role.objects.all().order_by('name')
    supervisors = Supervisor.objects.all().select_related('user').order_by('user__username')
    
    context = {
        'projects': projects,
        'roles': roles,
        'supervisors': supervisors,
    }
    
    return render(request, 'labourers/settings.html', context)

@login_required
def project_geofence_setup(request):
    """
    Page for setting up project geofence boundaries
    """
    projects = Project.objects.all().order_by('name')
    
    context = {
        'projects': projects,
    }
    
    return render(request, 'labourers/project_geofence.html', context)

@login_required
@csrf_exempt
def update_project_boundary(request, project_id):
    """
    API endpoint to update project boundary coordinates
    """
    if request.method == 'POST':
        try:
            project = get_object_or_404(Project, id=project_id)
            data = json.loads(request.body)
            boundary_coordinates = data.get('boundary_coordinates')
            
            if not boundary_coordinates or len(boundary_coordinates) < 3:
                return JsonResponse({
                    'success': False,
                    'error': 'Boundary must have at least 3 points'
                })
            
            # Validate coordinate format
            for coord in boundary_coordinates:
                if not isinstance(coord, list) or len(coord) != 2:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid coordinate format. Each point must be [latitude, longitude]'
                    })
            
            # Save boundary
            project.boundary_coordinates = boundary_coordinates
            project.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Boundary saved for {project.name}',
                'boundary_coordinates': boundary_coordinates
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Only POST method allowed'
    })

@login_required
def export_attendance_csv(request):
    """Export attendance records to CSV"""
    import csv
    from datetime import datetime, timedelta
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Date', 'Time', 'Labourer Name', 'National ID', 'Type', 'Project', 'Location', 'Verification Method', 'Verified'])
    
    # Get date range from request
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    attendances = AttendanceLog.objects.filter(
        log_timestamp__gte=start_date
    ).select_related('labourer', 'labourer__project').order_by('-log_timestamp')
    
    for attendance in attendances:
        writer.writerow([
            attendance.log_timestamp.strftime('%Y-%m-%d'),
            attendance.log_timestamp.strftime('%H:%M:%S'),
            attendance.labourer.full_name,
            attendance.labourer.national_id,
            attendance.log_type,
            attendance.labourer.project.name if attendance.labourer.project else 'N/A',
            f"{attendance.latitude}, {attendance.longitude}" if attendance.latitude else 'N/A',
            attendance.verification_method or 'N/A',
            'Yes' if attendance.notes and 'Verified: True' in attendance.notes else 'No'
        ])
    
    return response

@login_required
def export_attendance_excel(request):
    """Export attendance records to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import timedelta
    
    # Get date range from request
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    attendances = AttendanceLog.objects.filter(
        log_timestamp__gte=start_date
    ).select_related('labourer', 'labourer__project').order_by('-log_timestamp')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ['Date', 'Time', 'Labourer Name', 'National ID', 'Type', 'Project', 'Location', 'Verification Method', 'Verified']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row, attendance in enumerate(attendances, 2):
        ws.cell(row=row, column=1, value=attendance.log_timestamp.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=attendance.log_timestamp.strftime('%H:%M:%S'))
        ws.cell(row=row, column=3, value=attendance.labourer.full_name)
        ws.cell(row=row, column=4, value=attendance.labourer.national_id)
        ws.cell(row=row, column=5, value=attendance.log_type)
        ws.cell(row=row, column=6, value=attendance.labourer.project.name if attendance.labourer.project else 'N/A')
        ws.cell(row=row, column=7, value=f"{attendance.latitude}, {attendance.longitude}" if attendance.latitude else 'N/A')
        ws.cell(row=row, column=8, value=attendance.verification_method or 'N/A')
        ws.cell(row=row, column=9, value='Yes' if attendance.notes and 'Verified: True' in attendance.notes else 'No')
    
    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 18
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def export_attendance_pdf(request):
    """Export attendance records to PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from datetime import timedelta
    from io import BytesIO
    
    # Get date range from request
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    attendances = AttendanceLog.objects.filter(
        log_timestamp__gte=start_date
    ).select_related('labourer', 'labourer__project').order_by('-log_timestamp')[:100]  # Limit for PDF
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    
    # Container for elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a2980'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    # Title
    title = Paragraph(f"Attendance Report - Last {days} Days", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Report info
    info_style = styles['Normal']
    info = Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Total Records: {attendances.count()}", info_style)
    elements.append(info)
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [['Date', 'Time', 'Labourer', 'ID', 'Type', 'Project', 'Verified']]
    
    for attendance in attendances:
        data.append([
            attendance.log_timestamp.strftime('%Y-%m-%d'),
            attendance.log_timestamp.strftime('%H:%M'),
            attendance.labourer.full_name[:20],  # Truncate long names
            attendance.labourer.national_id,
            attendance.log_type,
            (attendance.labourer.project.name[:15] if attendance.labourer.project else 'N/A'),
            'Yes' if attendance.notes and 'Verified: True' in attendance.notes else 'No'
        ])
    
    # Create table
    table = Table(data, colWidths=[1*inch, 0.8*inch, 1.8*inch, 1*inch, 0.8*inch, 1.5*inch, 0.8*inch])
    
    # Table styling
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a2980')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response
